from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from contextlib2 import closing
from datetime import date

def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def add_values_for_indicators(file_name, value1, value2, value3):
    with closing(load_workbook(filename=file_name)) as wb:
        today = date.today()
        print("Today's date:", today)
        sheet_to_focus = 'Indici'

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == sheet_to_focus:
                break
        wb.active = s
        ws = wb.active
        tb = ws.tables["Table1"]
        coord = list(range_boundaries(tb.ref))

        data = [[today, value1, value2, value3]]
        for row in data:
            coord[-1] += 1 
            ws.append(row)

        tb.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"
        wb.save(file_name)