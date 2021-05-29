from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from contextlib2 import closing
from datetime import date

def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def add_value_for_second(file_name, value1, value2, value3, value4, value5, value6, value7, value8, value9, value10, value11, value12, value13, value14, value15, value16, value17, value18, value19, value20, value21, value22, value23, value24, value25, value26, value27, value28, value29, value30, value31, value32, value33, value34, value35, value36, value37, value38, value39):
    with closing(load_workbook(filename=file_name)) as wb:
        today = date.today()
        print("Today's date:", today)
        sheet_to_focus = '440000'

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == sheet_to_focus:
                break
        wb.active = s
        ws = wb.active
        tb = ws.tables["Table4"]

        coord = list(range_boundaries(tb.ref))

        data = [[today, value1, value2, value3, value4, value5, value6, value7, value8, value9, value10, value11, value12, value13, value14, value15, value16, value17, value18, value19, value20, value21, value22, value23, value24, value25, value26, value27, value28, value29, value30, value31, value32, value33, value34, value35, value36, value37, value38, value39]]
        for row in data:
            coord[-1] += 1 
            ws.append(row)

        tb.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"
        wb.save(file_name)